# -*- coding: utf-8 -*-
"""원본(표준데이터 JSON, KESS xlsx) → 정돈된 중간 테이블.

산출(메모리 반환; match_schools.py 가 소비):
  parse_std()  -> list[dict]  기관 메타 (학교명·campus·school_type·found_type·sido·addr)
  parse_kess() -> dict{year: list[dict]}  연도별 학교 (학교명·학제·campus·sido·found·enroll·faculty)

정규화 없이 '원본에 충실한' 필드만 담는다. 학교명 정규화/매칭은 match_schools.py 담당.
"""
import json
import sys
import warnings

warnings.filterwarnings("ignore")

try:
    from etl.config import METADATA_DIR, YEARS
except Exception:
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from etl.config import METADATA_DIR, YEARS

import openpyxl

RAW_DIR = METADATA_DIR / "raw"

# ── 시도 정규화: 다양한 표기 → REGION_MAP 이 쓰는 짧은 형태 ──────────────
SIDO_CANON = {
    "서울": "서울", "서울특별시": "서울",
    "부산": "부산", "부산광역시": "부산",
    "대구": "대구", "대구광역시": "대구",
    "인천": "인천", "인천광역시": "인천",
    "광주": "광주", "광주광역시": "광주",
    "대전": "대전", "대전광역시": "대전",
    "울산": "울산", "울산광역시": "울산",
    "세종": "세종", "세종특별자치시": "세종", "세종시": "세종",
    "경기": "경기", "경기도": "경기",
    "강원": "강원", "강원도": "강원", "강원특별자치도": "강원",
    "충북": "충북", "충청북도": "충북",
    "충남": "충남", "충청남도": "충남",
    "전북": "전북", "전라북도": "전북", "전북특별자치도": "전북",
    "전남": "전남", "전라남도": "전남",
    "경북": "경북", "경상북도": "경북",
    "경남": "경남", "경상남도": "경남",
    "제주": "제주", "제주특별자치도": "제주", "제주도": "제주",
}


def norm_sido(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s in SIDO_CANON:
        return SIDO_CANON[s]
    # 주소 앞 2글자로 재시도
    return SIDO_CANON.get(s[:2], s[:2] if s else "")


# ── 학교구분(학제/UNIV_SE) → 정규화 school_type ──────────────────────────
def norm_school_type(univ_se=None, schl_se=None, hakje=None, grad_gubun=None):
    """대학 / 전문대학 / 사이버대학 / 대학원대학 / 방송통신대학 등으로 정규화."""
    tokens = " ".join([str(x) for x in (univ_se, schl_se, hakje, grad_gubun) if x])
    if grad_gubun and "대학원대학" in str(grad_gubun):
        return "대학원대학"
    if "대학원대학" in tokens:
        return "대학원대학"
    if "사이버" in tokens or "원격" in tokens:
        return "사이버대학"
    if "방송통신" in tokens:
        return "방송통신대학"
    if "전문" in tokens or "기능대학" in tokens or "전공대학" in tokens:
        # '전문대학원'은 대학원(위에서 안 걸림)이지만 base 로는 잘 안 옴
        if "대학원" in tokens and "전문대학원" in tokens:
            return "대학원대학"
        return "전문대학"
    if hakje and "대학원" in str(hakje):
        return "대학원대학"
    return "대학"


def norm_campus(v):
    if v is None:
        return ""
    s = str(v).strip()
    if "분교" in s and "본" not in s:
        return "분교"
    if "본교" in s:
        return "본교"
    if "제2" in s or "제3" in s or "제4" in s or "캠퍼" in s:
        return s.replace("(", "").replace(")", "")
    return s


def parse_std():
    """표준데이터 → 기관 메타 레코드 목록."""
    recs = json.loads((RAW_DIR / "std_univ.json").read_text(encoding="utf-8"))
    out = []
    for r in recs:
        out.append({
            "name": (r.get("SCHL_NM") or "").strip(),
            "univ_se": (r.get("UNIV_SE_NM") or "").strip(),      # 대학/전문대학/대학원
            "schl_se": (r.get("SCHL_SE_NM") or "").strip(),      # 대학교/전문대학/사이버대학(대학)...
            "found_type": (r.get("FNDN_FORM_SE_NM") or "").strip(),
            "sido": norm_sido(r.get("CTPV_NM")),
            "campus": norm_campus(r.get("MAINBRANCH_NM")),
            "addr": (r.get("LCTN_ROAD_NM_ADDR") or r.get("LCTN_LOTNO_ADDR") or "").strip(),
        })
    return out


def _find_header(ws):
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        rows.append(r)
        vals = [str(c).strip() if c is not None else "" for c in r]
        if "학교명" in vals and "연도" in vals:
            return i, r, rows
        if i > 30:
            break
    raise RuntimeError("KESS 헤더 행을 찾지 못함")


def parse_kess():
    """KESS 연도별 xlsx → {year: [기관 레코드,...]} (base 기관만; 부설대학원 제외)."""
    out = {}
    for year in YEARS:
        path = RAW_DIR / ("kess_%d.xlsx" % year)
        if not path.exists():
            out[year] = []
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        hrow, hdr, _ = _find_header(ws)

        def col(label):
            for j, c in enumerate(hdr):
                if c is not None and str(c).strip() == label:
                    return j
            return None

        ci = {
            "name": col("학교명"), "hakje": col("학제"), "grad": col("대학원구분"),
            "state": col("학교상태"), "campus": col("본분교"), "sido": col("시도"),
            "found": col("설립"), "enroll": col("재학생_전체_계"),
            "enrolled_reg": col("재적생_전체_계"), "faculty": col("전임교원_계"),
        }
        recs = []
        for r in ws.iter_rows(min_row=hrow + 2, values_only=True):
            name = r[ci["name"]] if ci["name"] is not None else None
            if not name:
                continue
            grad = r[ci["grad"]] if ci["grad"] is not None else None
            if grad is not None and "부설대학원" in str(grad):
                continue  # 부설대학원 제외(재정계산서 대상은 본교 기관)
            def gv(k):
                j = ci[k]
                return r[j] if (j is not None and j < len(r)) else None
            def gi(k):
                v = gv(k)
                try:
                    return int(v)
                except (TypeError, ValueError):
                    return None
            recs.append({
                "name": str(name).strip(),
                "hakje": (str(gv("hakje")).strip() if gv("hakje") else ""),
                "grad": (str(grad).strip() if grad else ""),
                "state": (str(gv("state")).strip() if gv("state") else ""),
                "campus": norm_campus(gv("campus")),
                "sido": norm_sido(gv("sido")),
                "found": (str(gv("found")).strip() if gv("found") else ""),
                "enroll": gi("enroll"),
                "faculty": gi("faculty"),
            })
        out[year] = recs
    return out


if __name__ == "__main__":
    std = parse_std()
    kess = parse_kess()
    print("[parse] std records: %d" % len(std))
    for y in YEARS:
        print("[parse] kess %d base institutions: %d" % (y, len(kess.get(y, []))))
