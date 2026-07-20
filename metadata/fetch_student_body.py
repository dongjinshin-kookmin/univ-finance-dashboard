# -*- coding: utf-8 -*-
"""재적·재학·휴학 학생 스톡 수집 (D1) — 코호트 모델용.

소스: 이미 캐시된 KESS 학교별 xlsx(metadata/raw/kess_2016..2024.xlsx),
  시트 `학교별 교육통계`(헤더 14행, 134개 컬럼).
  base 기관 행(부설대학원 제외)에서 학생 스톡을 뽑고, 다캠퍼스는 campus_key 로
  묶어 **연도별 합산**한다. metadata/fetch_enrollment.py 와 동일한 매칭 계약을 재사용.

추출 컬럼(전부 `_전체_계` = base 대학교 행 기준 → 학부 스톡. 부설대학원 행을 빼므로
  석사/박사는 자연히 제외되고 남는 `전체`는 학부만 집계된다):
  · enrolled_total  = 재적생_전체_계  (재적 = 재학 + 휴학 + 유예)
  · attending_total = 재학생_전체_계  (재학)
  · on_leave        = 휴학생_전체_계  (휴학)
  · deferred        = 유예생_전체_계  (졸업유예; 항등식 재적=재학+휴학+유예 확인용)

학년별(1~4학년) 인원: KESS 「학교별 교육통계」 134개 컬럼에는 **존재하지 않는다.**
  분해는 학위과정(학사=전체 / 석사 / 박사)뿐이며 학년(year-of-study) 축이 없다.
  편입학 컬럼도 이 데이터셋에는 없다. 따라서 grade1..grade6 컬럼은 산출하지 않는다.
  (근거·전수 컬럼 목록은 build/d1_report.md 참조.)

산출: build/interim/student_body.csv
  canonical, year, enrolled_total, attending_total, on_leave, deferred
  build/d1_report.md (컬럼 목록·학년별 부재·검증·국민대 스팟체크)

매칭 계약: metadata/match_schools.py 의 normalize/clean/campus_key 재사용.
  개명 학교는 canonical+aliases 의 campus_key 를 유니온해 합산.
멱등: KESS 캐시를 읽기만 하므로 반복 실행 안전. --force 는 리포트 재작성만 강제.
"""
import csv
import sys
import warnings

warnings.filterwarnings("ignore")

try:
    from etl.config import INTERIM_DIR, METADATA_DIR, BUILD_DIR, YEARS, KMU_NAME
except Exception:
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from etl.config import INTERIM_DIR, METADATA_DIR, BUILD_DIR, YEARS, KMU_NAME

import openpyxl

from metadata.match_schools import normalize, clean, campus_key
from metadata import build_school_master as bsm

RAW_DIR = METADATA_DIR / "raw"
OUT_CSV = INTERIM_DIR / "student_body.csv"
REPORT = BUILD_DIR / "d1_report.md"

# KESS 「학교별 교육통계」 컬럼명 → 산출 필드
STOCK_COLS = {
    "enrolled_total": "재적생_전체_계",
    "attending_total": "재학생_전체_계",
    "on_leave": "휴학생_전체_계",
    "deferred": "유예생_전체_계",
}
FIELDS = list(STOCK_COLS.keys())


# ── KESS 파싱 → campus_key 단위 학생 스톡 집계 ────────────────────────────
def _find_header(ws):
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(c).strip() if c is not None else "" for c in r]
        if "학교명" in vals and "연도" in vals:
            return i, r
        if i > 40:
            break
    raise RuntimeError("KESS 헤더 행을 찾지 못함")


def build_kess_stock():
    """반환:
      ck2years : {campus_key: {year: {field: int|None, ...}}}
      norm2ck  : {정규화명: campus_key}
      clean2ck : {clean명: campus_key}
    """
    ck2years = {}
    norm2ck = {}
    clean2ck = {}
    for year in YEARS:
        path = RAW_DIR / ("kess_%d.xlsx" % year)
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        hi, hdr = _find_header(ws)
        col = {str(c).strip().replace("\n", ""): j for j, c in enumerate(hdr) if c}
        cn = col.get("학교명")
        cg = col.get("대학원구분")
        cfields = {f: col.get(src) for f, src in STOCK_COLS.items()}
        if cn is None:
            continue
        for r in ws.iter_rows(min_row=hi + 2, values_only=True):
            name = r[cn] if cn < len(r) else None
            if not name:
                continue
            grad = str(r[cg]).strip() if (cg is not None and cg < len(r) and r[cg]) else ""
            if "부설대학원" in grad:  # 대학의 부설대학원 행 제외(match_schools 계약과 동일)
                continue

            def gi(j):
                if j is None or j >= len(r):
                    return None
                try:
                    return int(r[j])
                except (TypeError, ValueError):
                    return None

            ck = campus_key(str(name))
            if not ck:
                continue
            c, _a = normalize(str(name))
            norm2ck.setdefault(c, ck)
            clean2ck.setdefault(clean(str(name)), ck)

            slot = ck2years.setdefault(ck, {}).setdefault(
                year, {f: None for f in FIELDS})
            for f, j in cfields.items():
                v = gi(j)
                if v is not None:
                    slot[f] = (slot[f] or 0) + v
    return ck2years, norm2ck, clean2ck


def _ck_set(candidates, ck2years, norm2ck, clean2ck):
    """canonical+aliases → 매칭되는 campus_key 집합(개명 유니온)."""
    cks = set()
    for n in candidates:
        c, _a = normalize(n)
        if c in norm2ck:
            cks.add(norm2ck[c])
            continue
        cc = clean(c)
        if cc in clean2ck:
            cks.add(clean2ck[cc])
            continue
        ck = campus_key(n)
        if ck in ck2years:
            cks.add(ck)
    return cks


# ── 메인 ────────────────────────────────────────────────────────────────
def main(force=False):
    targets = bsm.load_targets()
    ck2years, norm2ck, clean2ck = build_kess_stock()

    rows = []
    matched = 0
    full9 = 0
    unmatched = []
    kmu_dbg = {}
    identity_ok = 0
    identity_bad = 0
    ge_ok = 0   # 재적 >= 재학
    ge_bad = 0
    for canonical, aliases in targets:
        cks = _ck_set([canonical] + aliases, ck2years, norm2ck, clean2ck)
        if not cks:
            unmatched.append(canonical)
            continue
        matched += 1
        have_years = 0
        for year in YEARS:
            agg = {f: None for f in FIELDS}
            for ck in cks:
                yslot = ck2years.get(ck, {}).get(year)
                if not yslot:
                    continue
                for f in FIELDS:
                    if yslot[f] is not None:
                        agg[f] = (agg[f] or 0) + yslot[f]
            if any(agg[f] is not None for f in FIELDS):
                have_years += 1
            # 검증 지표(값이 있는 행만 집계)
            en, at, lv, df = (agg["enrolled_total"], agg["attending_total"],
                              agg["on_leave"], agg["deferred"])
            if en is not None and at is not None:
                if en >= at:
                    ge_ok += 1
                else:
                    ge_bad += 1
                if lv is not None and df is not None:
                    if en == at + lv + df:
                        identity_ok += 1
                    else:
                        identity_bad += 1
            rows.append({
                "canonical": canonical,
                "year": year,
                "enrolled_total": en if en is not None else "",
                "attending_total": at if at is not None else "",
                "on_leave": lv if lv is not None else "",
                "deferred": df if df is not None else "",
            })
            if canonical == KMU_NAME:
                kmu_dbg[year] = (en, at, lv, df)
        if have_years == len(YEARS):
            full9 += 1

    write_csv(rows)
    print("[sbody] matched %d/%d  (9개년 완비 %d)  rows=%d"
          % (matched, len(targets), full9, len(rows)))
    if unmatched:
        print("[sbody] unmatched: %s" % ", ".join(unmatched))
    print("[sbody][spot] %s (재적/재학/휴학/유예): %s" % (KMU_NAME, kmu_dbg))
    print("[sbody] 재적>=재학: ok=%d bad=%d | 재적=재학+휴학+유예: ok=%d bad=%d"
          % (ge_ok, ge_bad, identity_ok, identity_bad))

    stats = {
        "n_targets": len(targets),
        "matched": matched,
        "full9": full9,
        "unmatched": unmatched,
        "kmu": kmu_dbg,
        "n_rows": len(rows),
        "ge_ok": ge_ok, "ge_bad": ge_bad,
        "identity_ok": identity_ok, "identity_bad": identity_bad,
    }
    write_report(stats)
    return stats


def write_csv(rows):
    cols = ["canonical", "year"] + FIELDS
    INTERIM_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print("[sbody] wrote %s (%d rows)" % (OUT_CSV, len(rows)))


def write_report(s):
    L = []
    ap = L.append
    ap("# D1 데이터 수집 리포트 — 재적·재학·휴학 학생 스톡(코호트 모델용)\n")
    ap("입학정원 감소 시뮬레이션의 코호트 모델 정확도를 위해, 캐시된 KESS 원본에서 학교별 "
       "학생 스톡을 추출했다. 자금계산서 대상 대학 %d교 대상.\n" % s["n_targets"])

    ap("## 1. 소스 · 방법\n")
    ap("- 소스: metadata/raw/kess_2016..2024.xlsx, 시트 `학교별 교육통계`(헤더 14행, 134개 컬럼).")
    ap("- base 기관 행만 사용(**부설대학원 행 제외**), 다캠퍼스는 campus_key 로 묶어 연도별 합산.")
    ap("- 매칭 계약은 metadata/match_schools.py 의 normalize/clean/campus_key 재사용")
    ap("  (fetch_enrollment.py 와 동일). 개명 학교는 campus_key 유니온 합산.")
    ap("- 멱등: KESS 캐시를 읽기만 함. 반복 실행 안전.")
    ap("")

    ap("## 2. 확인한 학생 관련 컬럼(134개 전수 확인 결과)\n")
    ap("학생 스톡·유량 관련 컬럼은 모두 **학위과정(전체=학사 / 석사 / 박사) × 성별(계/남/여)** "
       "축으로만 분해되어 있고, **학년(1~4학년) 축과 편입학 컬럼은 존재하지 않는다.**\n")
    ap("| 유형 | 컬럼(전체_계 기준) | 산출 필드 |")
    ap("|---|---|---|")
    ap("| 재적생 | 재적생_전체_계 (석사/박사 각각 별도) | enrolled_total |")
    ap("| 재학생 | 재학생_전체_계 (석사/박사 각각 별도) | attending_total |")
    ap("| 휴학생 | 휴학생_전체_계 (석사/박사 각각 별도) | on_leave |")
    ap("| 유예생 | 유예생_전체_계 | deferred |")
    ap("| (입학정원) | 입학정원_전체 / 정원내_입학자_전체_계 | (fetch_enrollment.py 담당) |")
    ap("| (지원·입학) | 지원자·입학자·정원내/외 입학자 (전체/석사/박사) | 미사용 |")
    ap("| (외국인유학생) | 외국인유학생 총계/학위과정/학사/석사/박사/연수 | 미사용 |")
    ap("| (졸업·교원) | 졸업자·전임/비전임교원·시간강사·직원 | 미사용 |")
    ap("")
    ap("### 2-1. 학년별(cohort) 데이터 존재 여부 — **없음**\n")
    ap("- KESS 「학교별 교육통계」에는 1~4학년별 재적/재학 인원 컬럼이 **없다.** 분해축은 "
       "학위과정(학사/석사/박사)과 성별뿐이다.")
    ap("- 편입학(편입생) 컬럼도 이 데이터셋에는 없다.")
    ap("- 따라서 grade1..grade6 컬럼은 산출하지 않았다. 코호트 모델에서 학년 분포가 필요하면 "
       "별도 소스(대학알리미 학년별 재학생 현황 등)가 필요하다 — 이 데이터셋으로는 불가.")
    ap("")

    ap("## 3. 산출물 — build/interim/student_body.csv\n")
    ap("- 컬럼: canonical, year, enrolled_total, attending_total, on_leave, deferred")
    ap("- 매칭 **%d/%d교**, 9개년 전부 확보 **%d교**, 총 **%d행** (연도 2016~2024).\n"
       % (s["matched"], s["n_targets"], s["full9"], s["n_rows"]))
    if s["unmatched"]:
        ap("- 미매칭 %d교: %s" % (len(s["unmatched"]), ", ".join(s["unmatched"])))
    ap("- base 대학교 행 기준이므로 enrolled/attending 은 **학부 스톡**(석사/박사 제외). "
       "이는 schools_meta.csv 의 재학생 정의와 동일하다.")
    ap("")

    ap("## 4. 검증\n")
    ap("### (a) 국민대학교 2024 재학생 정합 — schools_meta.csv 교차검증\n")
    kmu24 = s["kmu"].get(2024)
    if kmu24:
        en, at, lv, df = kmu24
        ap("- student_body 국민대 2024: 재적 %s / **재학 %s** / 휴학 %s / 유예 %s"
           % (en, at, lv, df))
        ap("- schools_meta.csv enroll_2024 = 14,486 → **재학 %s 과 %s.**"
           % (at, "일치" if at == 14486 else "불일치"))
    ap("")
    ap("### 국민대 9개년 스팟체크 (재적 / 재학 / 휴학 / 유예)\n")
    for y in YEARS:
        if y in s["kmu"]:
            en, at, lv, df = s["kmu"][y]
            ap("  - %d: %s / %s / %s / %s" % (y, en, at, lv, df))
    ap("")
    ap("### (b) 커버리지 — enrollment_metrics.csv 대비\n")
    ap("- student_body 매칭 %d교 / enrollment_metrics 와 동일 targets·동일 매칭 로직이므로 "
       "커버리지 동일(미매칭 목록 위와 같음)." % s["matched"])
    ap("")
    ap("### (c) 재적 ≥ 재학 항등식\n")
    ap("- 재적 ≥ 재학: 정상 %d행 / 위반 %d행" % (s["ge_ok"], s["ge_bad"]))
    ap("- 강항등식 재적 = 재학 + 휴학 + 유예: 성립 %d행 / 불성립 %d행"
       % (s["identity_ok"], s["identity_bad"]))
    ap("")

    ap("## 5. 재현\n")
    ap("- 스크립트: metadata/fetch_student_body.py (멱등, --force 는 리포트 재작성만 강제)")
    ap("- 실행: `python3 -m metadata.fetch_student_body`")

    BUILD_DIR.mkdir(parents=True, exist_ok=True)
    REPORT.write_text("\n".join(L), encoding="utf-8")
    print("[sbody] wrote %s" % REPORT)


if __name__ == "__main__":
    main(force="--force" in sys.argv)
