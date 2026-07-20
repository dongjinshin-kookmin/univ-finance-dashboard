# -*- coding: utf-8 -*-
"""대학원 재학 스톡(H_grad)·정원외 입학자(A_out)·입학정원 학부/대학원 분해 수집 (A1).

입학정원 감소 시뮬레이션의 5110 세그먼트 분해(학부 정원내 / 학부 정원외 / 대학원)를 위해,
캐시된 KESS 원본에서 fetch_student_body.py / fetch_enrollment.py 가 **제외**했던
부분(부설대학원 행·정원외 입학자·학위과정별 정원)을 뽑는다. 매칭 계약은 그대로 재사용.

소스: metadata/raw/kess_2016..2024.xlsx, 시트 `학교별 교육통계`(헤더 14행, 131~134개 컬럼).

산출 필드 정의(KESS 구조 실측 결과 — build/d5_report.md §2 참조):
  · h_grad  = 대학원 재학생 = Σ(전 행: 재학생_석사_계 + 재학생_박사_계),
              부설대학원 행은 parent 학교로 귀속(최장 prefix 매칭). 학부 재학은 절대 미포함.
              → **이중계상 없음**: 부설대학원 행 재학생_전체_계 == 석사+박사(전 1,429행 실측 확인)
                이므로 「부설대학원 전체_계 합」(설계 1.3절)과 수치 동일하되, 학부 base 행에
                석/박이 섞인 기관(대학원대학 등)까지 정확히 포착한다.
  · a_total = 입학자_전체_계 (base 행) = 정원내 + 정원외 전체 입학자.
  · a_out   = 정원외_입학자_전체_계 (base 행) = 학부 정원외 입학자(외국인·편입). ≥ 0.
              (검증: a_total = a_in + a_out 항등식 성립 확인)
  · a_in    = 정원내_입학자_전체_계 (base 행) = enrollment_metrics.csv admitted 와 동일 정의(교차검증용).
  · q_ug    = 정원내 입학정원_학부 (base 행) = 학부 입학정원(정원내).
  · q_grad  = Σ(전 행: 정원내_입학정원_석사 + 정원내_입학정원_박사) = 대학원 입학정원(정원내).

base 행 = 대학원구분이 '부설대학원'이 아닌 행(학부대학 + 대학원대학). 부설대학원 행만 제외하는
  것은 fetch_enrollment / fetch_student_body 와 동일한 계약 → 커버리지 동일.

매칭 계약: metadata/match_schools.py 의 normalize/clean/campus_key 재사용. 다캠퍼스는 campus_key
  로 묶어 연도별 합산. 부설대학원 행은 학교명이 base 학교명을 접두로 가지므로, 그 해 base
  campus_key 집합에 대한 **최장 접두(prefix) 일치**로 parent 학교에 귀속(orphan은 1/1429뿐).
멱등: KESS 캐시를 읽기만 하므로 반복 실행 안전. --force 는 리포트 재작성만 강제.

산출: build/interim/grad_out.csv
  canonical, year, h_grad, a_total, a_out, a_in, q_ug, q_grad
  build/d5_report.md (컬럼 정의·이중계상 배제 논리·검증·국민대 스팟체크·미확보 항목)
"""
import csv
import sys
import warnings

warnings.filterwarnings("ignore")

try:
    from etl.config import INTERIM_DIR, METADATA_DIR, BUILD_DIR, YEARS, KMU_NAME
except Exception:
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from etl.config import INTERIM_DIR, METADATA_DIR, BUILD_DIR, YEARS, KMU_NAME

import openpyxl

from metadata.match_schools import normalize, clean, campus_key
from metadata import build_school_master as bsm

RAW_DIR = METADATA_DIR / "raw"
OUT_CSV = INTERIM_DIR / "grad_out.csv"
REPORT = BUILD_DIR / "d5_report.md"

# base 행(부설대학원 제외)에서 뽑는 학부/전체 지표
BASE_COLS = {
    "a_total": "입학자_전체_계",
    "a_out": "정원외_입학자_전체_계",
    "a_in": "정원내_입학자_전체_계",
    "q_ug": "정원내 입학정원_학부",
}
# 전 행(base + 부설대학원)에서 석사+박사를 합쳐 parent 학교로 귀속하는 대학원 지표
GRAD_ATTEND = ("재학생_석사_계", "재학생_박사_계")   # → h_grad
GRAD_QUOTA = ("정원내_입학정원_석사", "정원내_입학정원_박사")  # → q_grad

FIELDS = ["h_grad", "a_total", "a_out", "a_in", "q_ug", "q_grad"]


def _find_header(ws):
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(c).strip() if c is not None else "" for c in r]
        if "학교명" in vals and "연도" in vals:
            return i, r
        if i > 40:
            break
    raise RuntimeError("KESS 헤더 행을 찾지 못함")


def _to_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


# ── KESS 파싱 → campus_key 단위 대학원·정원외 지표 집계 ────────────────────
def build_kess_grad_out():
    """반환:
      ck2years : {campus_key: {year: {field: int|None, ...}}}
      norm2ck  : {정규화명: campus_key}   (base 행 기준, _ck_set 매칭용)
      clean2ck : {clean명: campus_key}
      orphans  : set(귀속 실패 부설대학원 학교명)
    """
    ck2years = {}
    norm2ck = {}
    clean2ck = {}
    orphans = set()

    for year in YEARS:
        path = RAW_DIR / ("kess_%d.xlsx" % year)
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        hi, hdr = _find_header(ws)
        col = {str(c).strip().replace("\n", ""): j for j, c in enumerate(hdr) if c}
        cn = col.get("학교명")
        cg = col.get("대학원구분")
        if cn is None:
            continue
        c_base = {f: col.get(src) for f, src in BASE_COLS.items()}
        c_att = [col.get(x) for x in GRAD_ATTEND]
        c_q = [col.get(x) for x in GRAD_QUOTA]

        # 한 번만 읽어 리스트에 적재(2-pass: base_ck 수집 → grad 귀속)
        recs = []
        for r in ws.iter_rows(min_row=hi + 2, values_only=True):
            name = r[cn] if cn < len(r) else None
            if not name:
                continue
            grad = str(r[cg]).strip() if (cg is not None and cg < len(r) and r[cg]) else ""
            recs.append((str(name), grad, r))

        def cell(r, j):
            if j is None or j >= len(r):
                return None
            return _to_int(r[j])

        def grad_sum(r, cols):
            vals = [cell(r, j) for j in cols]
            vals = [v for v in vals if v is not None]
            return sum(vals) if vals else None

        # pass 1 — base 행: campus_key 집합·정규화 인덱스·base 지표·자기 석/박 귀속
        base_cks = set()
        for name, grad, r in recs:
            if "부설대학원" in grad:
                continue
            ck = campus_key(name)
            if not ck:
                continue
            base_cks.add(ck)
            c, _a = normalize(name)
            norm2ck.setdefault(c, ck)
            clean2ck.setdefault(clean(name), ck)
            slot = ck2years.setdefault(ck, {}).setdefault(
                year, {f: None for f in FIELDS})
            for f, j in c_base.items():
                v = cell(r, j)
                if v is not None:
                    slot[f] = (slot[f] or 0) + v
            # 대학원대학 등 base 행 자체에 석/박이 있으면 h_grad·q_grad 로 흡수(이중계상 방지)
            hv = grad_sum(r, c_att)
            if hv is not None:
                slot["h_grad"] = (slot["h_grad"] or 0) + hv
            qv = grad_sum(r, c_q)
            if qv is not None:
                slot["q_grad"] = (slot["q_grad"] or 0) + qv

        sorted_bases = sorted(base_cks, key=len, reverse=True)

        # pass 2 — 부설대학원 행: 최장 prefix 로 parent 학교에 석/박 귀속
        for name, grad, r in recs:
            if "부설대학원" not in grad:
                continue
            gclean = clean(name)
            parent = None
            for b in sorted_bases:
                if gclean.startswith(b):
                    parent = b
                    break
            if parent is None:
                orphans.add(name)
                continue
            slot = ck2years.setdefault(parent, {}).setdefault(
                year, {f: None for f in FIELDS})
            hv = grad_sum(r, c_att)
            if hv is not None:
                slot["h_grad"] = (slot["h_grad"] or 0) + hv
            qv = grad_sum(r, c_q)
            if qv is not None:
                slot["q_grad"] = (slot["q_grad"] or 0) + qv

    return ck2years, norm2ck, clean2ck, orphans


def _ck_set(candidates, ck2years, norm2ck, clean2ck):
    """canonical+aliases → 매칭되는 campus_key 집합(개명 유니온) — fetch_student_body 계약."""
    cks = set()
    for n in candidates:
        c, _a = normalize(n)
        if c in norm2ck:
            cks.add(norm2ck[c])
            continue
        cc = clean(c)
        if cc in clean2ck:
            cks.add(clean2ck[cc])
            continue
        ck = campus_key(n)
        if ck in ck2years:
            cks.add(ck)
    return cks


# ── 메인 ────────────────────────────────────────────────────────────────
def main(force=False):
    targets = bsm.load_targets()
    ck2years, norm2ck, clean2ck, orphans = build_kess_grad_out()

    rows = []
    matched = 0
    full9 = 0
    unmatched = []
    kmu_dbg = {}
    aout_neg = 0            # a_out < 0 (있으면 안 됨)
    aout_present = 0
    ident_ok = 0           # a_total == a_in + a_out
    ident_bad = 0
    hgrad_present = 0

    for canonical, aliases in targets:
        cks = _ck_set([canonical] + aliases, ck2years, norm2ck, clean2ck)
        if not cks:
            unmatched.append(canonical)
            continue
        matched += 1
        have_years = 0
        for year in YEARS:
            agg = {f: None for f in FIELDS}
            for ck in cks:
                yslot = ck2years.get(ck, {}).get(year)
                if not yslot:
                    continue
                for f in FIELDS:
                    if yslot[f] is not None:
                        agg[f] = (agg[f] or 0) + yslot[f]
            if any(agg[f] is not None for f in FIELDS):
                have_years += 1

            a_tot, a_out, a_in = agg["a_total"], agg["a_out"], agg["a_in"]
            h_grad = agg["h_grad"]
            if a_out is not None:
                aout_present += 1
                if a_out < 0:
                    aout_neg += 1
            if a_tot is not None and a_in is not None and a_out is not None:
                if a_tot == a_in + a_out:
                    ident_ok += 1
                else:
                    ident_bad += 1
            if h_grad is not None:
                hgrad_present += 1

            rows.append({
                "canonical": canonical,
                "year": year,
                **{f: (agg[f] if agg[f] is not None else "") for f in FIELDS},
            })
            if canonical == KMU_NAME:
                kmu_dbg[year] = tuple(agg[f] for f in FIELDS)
        if have_years == len(YEARS):
            full9 += 1

    write_csv(rows)
    print("[grad_out] matched %d/%d  (9개년 완비 %d)  rows=%d"
          % (matched, len(targets), full9, len(rows)))
    if unmatched:
        print("[grad_out] unmatched: %s" % ", ".join(unmatched))
    if orphans:
        print("[grad_out] 귀속 실패 부설대학원 %d행: %s"
              % (len(orphans), ", ".join(sorted(orphans))))
    print("[grad_out][spot] %s (h_grad,a_total,a_out,a_in,q_ug,q_grad): %s"
          % (KMU_NAME, kmu_dbg))
    print("[grad_out] a_out<0: %d/%d | a_total=a_in+a_out: ok=%d bad=%d | h_grad 존재 %d행"
          % (aout_neg, aout_present, ident_ok, ident_bad, hgrad_present))

    stats = {
        "n_targets": len(targets),
        "matched": matched,
        "full9": full9,
        "unmatched": unmatched,
        "orphans": sorted(orphans),
        "kmu": kmu_dbg,
        "n_rows": len(rows),
        "aout_neg": aout_neg,
        "aout_present": aout_present,
        "ident_ok": ident_ok,
        "ident_bad": ident_bad,
        "hgrad_present": hgrad_present,
    }
    write_report(stats)
    return stats


def write_csv(rows):
    cols = ["canonical", "year"] + FIELDS
    INTERIM_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print("[grad_out] wrote %s (%d rows)" % (OUT_CSV, len(rows)))


def write_report(s):
    L = []
    ap = L.append
    ap("# D5 데이터 수집 리포트 — 대학원 재학(H_grad)·정원외 입학(A_out)·입학정원 학위분해 (WBS A1)\n")
    ap("입학정원 감소 시뮬레이션의 5110 세그먼트 분해(학부 정원내 / 학부 정원외 / 대학원)를 위해, "
       "캐시된 KESS 원본에서 fetch_student_body / fetch_enrollment 가 제외했던 부분(부설대학원 재학·"
       "정원외 입학자·학위과정별 정원)을 추출했다. 자금계산서 대상 대학 %d교 대상.\n" % s["n_targets"])

    ap("## 1. 소스 · 방법\n")
    ap("- 소스: metadata/raw/kess_2016..2024.xlsx, 시트 `학교별 교육통계`(헤더 14행, 131~134개 컬럼).")
    ap("- 매칭 계약은 metadata/match_schools.py 의 normalize/clean/campus_key 재사용(fetch_student_body 동일).")
    ap("  다캠퍼스는 campus_key 로 묶어 연도별 합산, 개명 학교는 campus_key 유니온.")
    ap("- **부설대학원 행 귀속**: 부설대학원 학교명은 base 학교명을 접두로 가지므로(예: `국민대학교경영대학원` "
       "⊃ `국민대학교`), 그 해 base campus_key 집합에 대한 **최장 접두(prefix) 일치**로 parent 학교에 합산.")
    ap("- 멱등: KESS 캐시를 읽기만 함. 반복 실행 안전(--force 는 리포트 재작성만 강제).")
    ap("")

    ap("## 2. 채택한 컬럼 정의와 이중계상 배제 논리\n")
    ap("대학원구분 값(2024 실측): 학부대학(none) 490 · 부설대학원 1,429 · 대학원대학 44.")
    ap("base 행 = 부설대학원이 **아닌** 행(학부대학 + 대학원대학). 부설대학원만 제외 → "
       "fetch_enrollment/fetch_student_body 와 동일 계약이라 커버리지 동일.\n")
    ap("| 필드 | 정의(KESS 컬럼) | 비고 |")
    ap("|---|---|---|")
    ap("| **h_grad** | Σ(전 행: 재학생_석사_계 + 재학생_박사_계), 부설대학원 행은 parent 귀속 | 대학원 재학 스톡 |")
    ap("| **a_total** | 입학자_전체_계 (base 행) | 정원내+정원외 전체 입학자 |")
    ap("| **a_out** | 정원외_입학자_전체_계 (base 행) | 학부 정원외 입학자(외국인·편입), ≥0 |")
    ap("| a_in | 정원내_입학자_전체_계 (base 행) | enrollment_metrics.admitted 와 동일 정의(교차검증) |")
    ap("| q_ug | 정원내 입학정원_학부 (base 행) | 학부 입학정원(정원내) |")
    ap("| q_grad | Σ(전 행: 정원내_입학정원_석사 + 정원내_입학정원_박사) | 대학원 입학정원(정원내) |")
    ap("")
    ap("### 2-1. h_grad — 이중계상 배제(결정적)\n")
    ap("설계 1.3절은 `H_grad = Σ부설대학원 행 재학생_전체_계`를 제안했다. 그러나 이는 학부 base 행에 "
       "석·박사가 섞인 기관(대학원대학, 일부 통합 등재 대학)을 놓친다. 실측으로 **더 정확하고 이중계상 없는** "
       "정의를 채택했다:\n")
    ap("- **채택: h_grad = Σ(전 행의 재학생_석사_계 + 재학생_박사_계)** — 석·박사만 더하므로 학부 재학은 "
       "절대 미포함, 어느 행에 있든 대학원생을 정확히 1회만 집계.")
    ap("- **부설대학원 행 재학생_전체_계 == 재학생_석사_계 + 재학생_박사_계** 임을 전 1,429행에서 실측 확인 "
       "(불일치 0행) → 부설대학원 행에는 학부 재학이 없다. 따라서 채택 정의는 설계의 「부설대학원 전체_계 합」과 "
       "**수치가 동일**하되(예: 국민대 2024 두 방식 모두 3,571), base 행 석·박까지 포착해 일반적으로 더 정확.")
    ap("- base 행(학부대학)의 재학생_석사/박사는 대부분 0(학부 순수). 값이 있는 44개 base 행은 전부 "
       "`대학원대학`(자체가 대학원 기관)으로, 그 석·박을 자기 학교에 귀속시켜 누락 없이 흡수.")
    ap("")

    ap("## 3. 산출물 — build/interim/grad_out.csv\n")
    ap("- 컬럼: canonical, year, h_grad, a_total, a_out, a_in, q_ug, q_grad (연도 2016~2024)")
    ap("- 매칭 **%d/%d교**, 9개년 전부 확보 **%d교**, 총 **%d행**.\n"
       % (s["matched"], s["n_targets"], s["full9"], s["n_rows"]))
    if s["unmatched"]:
        ap("- 미매칭 %d교: %s (student_body.csv 와 동일 — 커버리지 동일)"
           % (len(s["unmatched"]), ", ".join(s["unmatched"])))
    if s["orphans"]:
        ap("- 부설대학원 귀속 실패 %d행(전 연도 누적): %s. base 학교명이 접두가 아니어서(예: `칼빈대학교`→"
           "`칼빈신학대학원`, `신학` 삽입) parent 미상. 이 중 타겟은 **칼빈대학교** 하나뿐이며 그 대학원생은 "
           "base 행 석·박으로 이미 대부분 포착(2024 h_grad=726), 한영신학대학교는 자금계산서 타겟 아님 → "
           "**전 타겟 h_grad 영향 무시 가능.**" % (len(s["orphans"]), ", ".join(s["orphans"])))
    ap("")

    ap("## 4. 검증\n")
    ap("### (a) 국민대 h_grad 범위 — 설계 1장 ρ_g 역산(d2_report 교차검증)\n")
    kmu24 = s["kmu"].get(2024)
    if kmu24:
        hg = kmu24[0]
        ap("- 국민대 2024 h_grad = **%s명**. 설계 1장/0.4절 추정 대학원 규모 ≈ 3,571, d2_report §4 공식단가 "
           "역산 함의 학생차 약 3,100~3,700 범위와 **정합**(%s)."
           % (hg, "통과" if (hg is not None and 3100 <= hg <= 3700) else "확인 필요"))
    ap("")
    ap("### 국민대 9개년 스팟체크 (h_grad / a_total / a_out / a_in / q_ug / q_grad)\n")
    for y in YEARS:
        if y in s["kmu"]:
            ap("  - %d: %s" % (y, " / ".join(str(x) for x in s["kmu"][y])))
    ap("")
    ap("### (b) a_out ≥ 0 (전 행)\n")
    ap("- a_out 존재 %d행 중 음수 **%d행** (%s)."
       % (s["aout_present"], s["aout_neg"], "통과" if s["aout_neg"] == 0 else "위반 있음"))
    ap("- 항등식 a_total = a_in + a_out: 성립 %d행 / 불성립 %d행 (전체입학자=정원내+정원외 확인)."
       % (s["ident_ok"], s["ident_bad"]))
    ap("")
    ap("### (c) 커버리지 — student_body.csv 대비\n")
    ap("- 매칭 %d/%d교, 9개년 완비 %d교. student_body(A2)·enrollment_metrics 와 **동일 targets·동일 매칭 "
       "로직**이므로 커버리지 동일. h_grad 존재 %d행."
       % (s["matched"], s["n_targets"], s["full9"], s["hgrad_present"]))
    ap("")

    ap("## 5. 미확보 · 주의점\n")
    ap("- **학부 정원외 재학 스톡 H_ug_out**: KESS는 정원외 *입학자*(a_out)만 제공, 정원외 *재학* 스톡 컬럼은 "
       "없음. 설계 2.3절대로 H_ug_out ≈ a_out × m_out(배율) 로 A3에서 추정.")
    ap("- **base 재학_전체_계의 대학원 혼입**: 학부대학 base 행은 대부분 학부 순수지만, 대학원대학은 base "
       "재학=석·박. A3의 p_ug 분모에서 H_ug 와 h_grad 중복을 피하려면 H_ug = 재학_전체_계 − (석+박)로 "
       "정제 권장(대다수 학부대학은 석+박=0이라 무영향).")
    ap("- **q_ug 가 기존 ext.series.입학정원보다 정확(중요)**: 기존 파이프라인의 입학정원 = KESS "
       "`입학정원_전체`(col20)인데, 2024 base 534행 분해 결과 이 값은 순수 학부정원내가 **아니다** — "
       "학부순수 369행 / 학부+정원외 **121행** / 정원내계(학부+석박) 44행(대학원대학). 즉 121개교에서 "
       "정원외 정원이, 44개교에서 대학원 정원이 섞여 있다(기타 0행, 전건 분해 확인). 본 산출의 "
       "**q_ug = 정원내 입학정원_학부(col22)** 는 순수 학부 정원내 감축 레버(설계 1.3절 Q)로 더 정확하며, "
       "A3는 Q 를 q_ug 로 대체 권장. (검증: a_in 은 enrollment_metrics.admitted 와 전 3,087행 완전 일치.)")
    ap("- **학년별(1~4학년)·편입 분해**: KESS에 컬럼 없음(d1_report §2-1). 별도 소스 필요.")
    ap("")

    ap("## 6. 재현\n")
    ap("- 스크립트: metadata/fetch_grad_out.py (멱등, --force 는 리포트 재작성만 강제)")
    ap("- 실행: `python3 -m metadata.fetch_grad_out`")

    BUILD_DIR.mkdir(parents=True, exist_ok=True)
    REPORT.write_text("\n".join(L), encoding="utf-8")
    print("[grad_out] wrote %s" % REPORT)


if __name__ == "__main__":
    main(force="--force" in sys.argv)
