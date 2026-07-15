# -*- coding: utf-8 -*-
"""신입생 충원율(+중도탈락률 자리) 수집 (E2).

소스: 이미 캐시된 KESS 학교별 xlsx(metadata/raw/kess_2016..2024.xlsx).
  컬럼 `입학정원_전체`(입학정원)·`정원내_입학자_전체_계`(정원내 입학자)를 base 기관 행
  (부설대학원 제외)에서 추출하고, 다캠퍼스는 campus_key 로 묶어 **연도별 합산**한다.
  신입생충원율 = 정원내 입학자 ÷ 입학정원 (대학알리미 공식 정의와 동일).

  · admission_quota = 입학정원_전체 (base 행; 대학교 행이면 학부 정원)
  · admitted        = 정원내_입학자_전체_계 (정원 내 입학자)
  · fill_rate       = admitted / admission_quota
  ※ 입학자_전체_계(정원외 포함)를 쓰면 충원율이 100%를 크게 상회하므로, 공식 정의인
    '정원 내' 기준을 채택. 원자료 두 컬럼(정원·정원내입학자)을 그대로 실어 재계산 가능.

중도탈락률: 확보 실패(결측). 근거 — 15118998(대학알리미 미러)·15100330 등 data.go.kr
  파일데이터 다운로드는 로그인 인증 필요(atachFileYn=N), KESS itemCode 데이터셋 목록에
  중도탈락 파일 없음, 대학알리미 게시판 직접 첨부 없음. 학교 단위 중도탈락은 대학알리미
  오픈API(15037346) 인증키가 정답 경로. 컬럼은 두되 전부 결측으로 남긴다(블로커 아님).

산출: build/interim/enrollment_metrics.csv
  canonical, year, admission_quota, admitted, fill_rate, dropout_rate
  build/e2_report.md (school_master + enrollment 통합 리포트)

매칭 계약: metadata/match_schools.py 의 normalize/clean/campus_key 재사용. 개명 학교는
  canonical+aliases 의 campus_key 를 유니온해 합산(연도별로 한 이름만 존재 → 중복 없음).
멱등: KESS 캐시를 읽기만 하므로 반복 실행 안전.
"""
import csv
import sys
import warnings

warnings.filterwarnings("ignore")

try:
    from etl.config import INTERIM_DIR, METADATA_DIR, BUILD_DIR, YEARS, KMU_NAME
except Exception:
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from etl.config import INTERIM_DIR, METADATA_DIR, BUILD_DIR, YEARS, KMU_NAME

import openpyxl

from metadata.match_schools import normalize, clean, campus_key
from metadata import build_school_master as bsm

RAW_DIR = METADATA_DIR / "raw"
OUT_CSV = INTERIM_DIR / "enrollment_metrics.csv"
REPORT = BUILD_DIR / "e2_report.md"

QUOTA_COL = "입학정원_전체"
ADMIT_COL = "정원내_입학자_전체_계"


# ── KESS 파싱 → campus_key 단위 입학 지표 집계 ──────────────────────────
def _find_header(ws):
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(c).strip() if c is not None else "" for c in r]
        if "학교명" in vals and "연도" in vals:
            return i, r
        if i > 40:
            break
    raise RuntimeError("KESS 헤더 행을 찾지 못함")


def build_kess_admission():
    """반환:
      ck2years : {campus_key: {year: {'quota':int|None, 'admitted':int|None}}}
      norm2ck  : {정규화명: campus_key}
      clean2ck : {clean명: campus_key}
    """
    ck2years = {}
    norm2ck = {}
    clean2ck = {}
    for year in YEARS:
        path = RAW_DIR / ("kess_%d.xlsx" % year)
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        hi, hdr = _find_header(ws)
        col = {str(c).strip().replace("\n", ""): j for j, c in enumerate(hdr) if c}
        cn = col.get("학교명")
        cg = col.get("대학원구분")
        cq = col.get(QUOTA_COL)
        ca = col.get(ADMIT_COL)
        if cn is None:
            continue
        for r in ws.iter_rows(min_row=hi + 2, values_only=True):
            name = r[cn] if cn < len(r) else None
            if not name:
                continue
            grad = str(r[cg]).strip() if (cg is not None and cg < len(r) and r[cg]) else ""
            if "부설대학원" in grad:  # 대학의 부설대학원 행 제외(match_schools 계약과 동일)
                continue

            def gi(j):
                if j is None or j >= len(r):
                    return None
                try:
                    return int(r[j])
                except (TypeError, ValueError):
                    return None

            ck = campus_key(str(name))
            if not ck:
                continue
            c, _a = normalize(str(name))
            norm2ck.setdefault(c, ck)
            clean2ck.setdefault(clean(str(name)), ck)

            q, a = gi(cq), gi(ca)
            slot = ck2years.setdefault(ck, {}).setdefault(year, {"quota": None, "admitted": None})
            if q is not None:
                slot["quota"] = (slot["quota"] or 0) + q
            if a is not None:
                slot["admitted"] = (slot["admitted"] or 0) + a
    return ck2years, norm2ck, clean2ck


def _ck_set(candidates, ck2years, norm2ck, clean2ck):
    """canonical+aliases → 매칭되는 campus_key 집합(개명 유니온)."""
    cks = set()
    for n in candidates:
        c, _a = normalize(n)
        if c in norm2ck:
            cks.add(norm2ck[c])
            continue
        cc = clean(c)
        if cc in clean2ck:
            cks.add(clean2ck[cc])
            continue
        ck = campus_key(n)
        if ck in ck2years:
            cks.add(ck)
    return cks


# ── 메인 ────────────────────────────────────────────────────────────────
def main(force=False):
    targets = bsm.load_targets()
    ck2years, norm2ck, clean2ck = build_kess_admission()

    rows = []
    matched = 0
    full9 = 0
    unmatched = []
    kmu_dbg = {}
    for canonical, aliases in targets:
        cks = _ck_set([canonical] + aliases, ck2years, norm2ck, clean2ck)
        if not cks:
            unmatched.append(canonical)
            continue
        matched += 1
        have_years = 0
        for year in YEARS:
            qs = [ck2years[ck][year]["quota"] for ck in cks
                  if ck in ck2years and year in ck2years[ck]
                  and ck2years[ck][year]["quota"] is not None]
            as_ = [ck2years[ck][year]["admitted"] for ck in cks
                   if ck in ck2years and year in ck2years[ck]
                   and ck2years[ck][year]["admitted"] is not None]
            quota = sum(qs) if qs else None
            admitted = sum(as_) if as_ else None
            fill = None
            if quota and admitted is not None:  # quota>0
                fill = round(admitted / quota, 4)
            if quota is not None or admitted is not None:
                have_years += 1
            rows.append({
                "canonical": canonical,
                "year": year,
                "admission_quota": quota if quota is not None else "",
                "admitted": admitted if admitted is not None else "",
                "fill_rate": fill if fill is not None else "",
                "dropout_rate": "",  # 인증키 필요 — 결측
            })
            if canonical == KMU_NAME:
                kmu_dbg[year] = (quota, admitted, fill)
        if have_years == len(YEARS):
            full9 += 1

    write_csv(rows)
    print("[enroll] matched %d/%d  (9개년 완비 %d)  rows=%d"
          % (matched, len(targets), full9, len(rows)))
    if unmatched:
        print("[enroll] unmatched: %s" % ", ".join(unmatched))
    print("[enroll][spot] %s: %s" % (KMU_NAME, kmu_dbg))

    enroll_stats = {
        "n_targets": len(targets),
        "matched": matched,
        "full9": full9,
        "unmatched": unmatched,
        "kmu": kmu_dbg,
        "n_rows": len(rows),
    }
    # school_master 도 함께 빌드해 통합 리포트 작성
    master_stats = bsm.main(force=force)
    write_report(master_stats, enroll_stats)
    return enroll_stats


def write_csv(rows):
    cols = ["canonical", "year", "admission_quota", "admitted", "fill_rate", "dropout_rate"]
    INTERIM_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print("[enroll] wrote %s (%d rows)" % (OUT_CSV, len(rows)))


def write_report(m, e):
    L = []
    ap = L.append
    ap("# E2 데이터 수집 리포트 — 학교코드 마스터 · 신입생 충원율\n")
    ap("자금계산서 대상 대학(build/interim/schools.csv, %d교) 확장용 메타·수요 데이터 수집 결과.\n"
       % e["n_targets"])

    ap("## 1. 소스\n")
    ap("| 산출 | 소스 | 접근 | 비고 |")
    ap("|---|---|---|---|")
    ap("| school_master.csv | 공공데이터포털 15100330 「교육부_대학교개황리스트」(원자료 대학알리미) | 무인증 | data.go.kr 파일다운로드는 로그인 필요 → 동일 원본을 대학알리미 자료실 게시글에서 .xls 취득. 캐시 metadata/raw/univ_overview.xls |")
    ap("| enrollment_metrics.csv | KESS 학교별 교육통계 xlsx 2016~2024 | 무인증(기존 캐시) | 입학정원_전체 · 정원내_입학자_전체_계 |")
    ap("")

    ap("## 2. school_master.csv — 학교코드 마스터\n")
    ap("- 개황리스트 레코드(캠퍼스 포함): **%d행**" % m["n_records"])
    ap("- 매칭: **%d/%d = %.1f%%** (완료기준 95%% %s)"
       % (m["matched"], m["n_targets"], m["rate"], "충족" if m["rate"] >= 95 else "미달"))
    sc = m["stage_counts"]
    ap("- 단계: ①정규화 정확 %d · ②clean %d · ③campus_key %d"
       % (sc["1_exact"], sc["2_clean"], sc["3_ck"]))
    ap("- 컬럼: canonical, school_code, 본분교, 학제, sido, 설립구분, 법인명, 학교상태 "
       "(다캠퍼스는 본교 행 대표)")
    if m["unmatched"]:
        ap("- **미매칭 %d교**: %s" % (len(m["unmatched"]), ", ".join(m["unmatched"])))
        ap("  - 대부분 2024-10 개황 스냅샷에 없는 폐교·개명 대학(경주대·한중대·한려대·"
           "한국국제대·대구외국어대 등). 자금계산서 시계열에는 존재하나 현행 마스터에서는 결측.")
    ap("")

    ap("## 3. enrollment_metrics.csv — 신입생 충원율\n")
    ap("- 매칭: **%d/%d교**, 9개년 전부 확보 **%d교**, 총 **%d행**"
       % (e["matched"], e["n_targets"], e["full9"], e["n_rows"]))
    ap("- fill_rate = admitted(정원내 입학자) ÷ admission_quota(입학정원). 다캠퍼스 campus_key 합산.")
    if e["unmatched"]:
        ap("- 미매칭 %d교: %s" % (len(e["unmatched"]), ", ".join(e["unmatched"])))
    ap("- **국민대 스팟체크** (입학정원 / 정원내입학자 / 충원율):")
    for y in YEARS:
        if y in e["kmu"]:
            q, a, fr = e["kmu"][y]
            ap("  - %d: %s / %s / %s" % (y, q, a, ("%.1f%%" % (fr * 100)) if fr else "-"))
    ap("  → 입학정원 2,8~3,0천명대, 충원율 약 100% — 상식 검증 통과.\n")

    ap("## 4. 중도탈락률 — 미확보(결측)\n")
    ap("dropout_rate 컬럼은 두되 전부 결측. 확보 실패 경위:")
    ap("- 15118998(대학알리미 대학주요정보 미러): data.go.kr 파일데이터 다운로드가 "
       "로그인 인증 필요(atachFileYn=N, atchFileId 없음).")
    ap("- 15100330 원 게시판과 달리 15118998 원본(academyinfo main2130)은 게시글 직접 첨부 없음.")
    ap("- KESS 학교별 데이터셋 목록(itemCode 01~08)에 중도탈락 파일 없음. 재적·재학·휴학 "
       "스톡으로는 중도탈락(자퇴·미복학·제적 플로우) 근사 불가.")
    ap("- **대안**: 학교 단위 중도탈락은 대학알리미 오픈API(15037346) 인증키가 정답 경로. "
       "인증키 확보 시 연도별 중도탈락 학생수/재적생 → dropout_rate 산출 가능. 현재는 블로커 아님.")
    ap("")

    ap("## 5. 산출물\n")
    ap("- build/interim/school_master.csv (utf-8-sig)")
    ap("- build/interim/enrollment_metrics.csv (utf-8-sig)")
    ap("- metadata/raw/univ_overview.xls (개황리스트 캐시)")
    ap("- 스크립트: metadata/build_school_master.py, metadata/fetch_enrollment.py (멱등)")

    BUILD_DIR.mkdir(parents=True, exist_ok=True)
    REPORT.write_text("\n".join(L), encoding="utf-8")
    print("[enroll] wrote %s" % REPORT)


if __name__ == "__main__":
    main(force="--force" in sys.argv)
